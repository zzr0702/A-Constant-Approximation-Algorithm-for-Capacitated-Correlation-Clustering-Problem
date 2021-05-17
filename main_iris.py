from JiSai.JiSai_Project_4 import function
from openpyxl import Workbook
import numpy as np

wb = Workbook()
ws = wb.active
N = list(np.arange(10, 30, 10))
PERCENTAGE = [0.15, 0.20, 0.25, 0.35]

for i in range(len(N)):
    ws.cell(2+i,1,N[i])

for p in range(len(PERCENTAGE)):
    ws.cell(1, p+2, 'U:' + str(PERCENTAGE[p]))
    for j in range(len(N)):
        U = int(PERCENTAGE[p] * N[j])
        sol = []
        for i in range(10):
            G = function.G_iris(N[j])
            solutionLP = function.solutionLP(G,U)
            solutionAlg = function.algorithm(G,U,solutionLP[0])
            if solutionLP[1] == 0:
                sol.append(1)
            else:
                sol.append(solutionAlg[1] / solutionLP[1])
        res = np.mean(sol)
        ws.cell(2+j,2+p,res)

wb.save('iris.xlsx')



